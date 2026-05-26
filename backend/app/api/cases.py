from io import BytesIO
import logging
import warnings

from fastapi import APIRouter, File, HTTPException, UploadFile
from openpyxl import load_workbook

from app.schemas.case import CaseCreate, CaseOut, CaseUpdate

router = APIRouter(prefix="/api/cases", tags=["cases"])
logger = logging.getLogger(__name__)

_cases_store: dict[int, CaseOut] = {}
_case_id_counter = 0


@router.get("", response_model=list[CaseOut])
def list_cases() -> list[CaseOut]:
    return list(_cases_store.values())


@router.post("", response_model=CaseOut, status_code=201)
def create_case(payload: CaseCreate) -> CaseOut:
    global _case_id_counter
    _case_id_counter += 1
    created = CaseOut(id=_case_id_counter, **payload.model_dump())
    _cases_store[created.id] = created
    return created


@router.put("/{case_id}", response_model=CaseOut)
def update_case(case_id: int, payload: CaseUpdate) -> CaseOut:
    existing = _cases_store.get(case_id)
    if existing is None:
        raise HTTPException(status_code=404, detail="Case not found")

    updated_data = existing.model_dump()
    updated_data.update(payload.model_dump(exclude_unset=True))
    updated = CaseOut(**updated_data)
    _cases_store[case_id] = updated
    return updated


@router.delete("/{case_id}", status_code=204)
def delete_case(case_id: int) -> None:
    if case_id not in _cases_store:
        raise HTTPException(status_code=404, detail="Case not found")
    _cases_store.pop(case_id)


@router.post("/import", response_model=list[CaseOut], status_code=201)
async def import_cases(file: UploadFile = File(...)) -> list[CaseOut]:
    global _case_id_counter

    logger.info("Import cases request received, filename=%s", file.filename)

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        logger.warning("Import rejected due to invalid file extension, filename=%s", file.filename)
        raise HTTPException(status_code=400, detail="Only .xlsx/.xlsm files are supported")

    payload = await file.read()
    logger.info("Excel upload payload read complete, filename=%s, size=%d bytes", file.filename, len(payload))

    with warnings.catch_warnings(record=True) as caught_warnings:
        warnings.simplefilter("always")
        workbook = load_workbook(filename=BytesIO(payload), read_only=True, data_only=True)
    for warning_info in caught_warnings:
        logger.warning(
            "Warning while loading workbook, filename=%s, category=%s, message=%s",
            file.filename,
            warning_info.category.__name__,
            str(warning_info.message),
        )
    sheet = workbook.active

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        workbook.close()
        logger.warning("Import failed because Excel file is empty, filename=%s", file.filename)
        raise HTTPException(status_code=400, detail="Excel file is empty")

    normalized_headers = [str(cell).strip().lower() if cell is not None else "" for cell in header_row]
    depth_aliases = {"depth", "层级", "层次"}
    fallback_offset = 1 if normalized_headers and normalized_headers[0] in depth_aliases else 0
    header_aliases: dict[str, tuple[str, ...]] = {
        "name": ("name", "名称", "用例名称"),
        "code": ("code", "编号", "用例编号"),
        "precondition": ("precondition", "前置条件", "预置条件"),
        "design_desc": ("设计描述", "设计说明", "description", "desc"),
        "steps": ("steps", "步骤", "测试步骤"),
        "expected": ("expected", "预期", "预期结果"),
    }
    normalized_aliases: dict[str, set[str]] = {
        field: {alias.strip().lower() for alias in aliases}
        for field, aliases in header_aliases.items()
    }
    header_index: dict[str, int] = {}
    for field, aliases in normalized_aliases.items():
        for i, header in enumerate(normalized_headers):
            if header and header in aliases:
                header_index[field] = i
                break
    logger.info(
        "Excel header parsed, filename=%s, headers=%s, detected_header_index=%s, fallback_offset=%d",
        file.filename,
        normalized_headers,
        header_index,
        fallback_offset,
    )

    def _get_cell(cells: tuple, field: str, fallback_index: int) -> str:
        index = header_index.get(field, fallback_index)
        if index >= len(cells) or cells[index] is None:
            return ""
        return str(cells[index]).strip()

    imported: list[CaseOut] = []
    skipped_rows_missing_name = 0
    skipped_rows_missing_required_fields = 0
    scanned_rows = 0
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row is None:
            continue
        scanned_rows += 1

        name = _get_cell(row, "name", fallback_offset)
        if not name:
            skipped_rows_missing_name += 1
            continue

        code = _get_cell(row, "code", fallback_offset + 1)
        precondition = _get_cell(row, "precondition", fallback_offset + 2)
        design_desc = _get_cell(row, "design_desc", fallback_offset + 3)
        steps = _get_cell(row, "steps", fallback_offset + 3)
        expected = _get_cell(row, "expected", fallback_offset + 4)

        if not steps and design_desc:
            steps = design_desc
        if not expected and design_desc:
            expected = design_desc

        if not steps or not expected:
            # CaseOut 要求 steps / expected 最少 1 个字符；缺失时跳过该行，避免 500
            skipped_rows_missing_required_fields += 1
            logger.debug(
                "Skip row due to missing required field, filename=%s, row=%d, has_steps=%s, has_expected=%s, has_design_desc=%s",
                file.filename,
                idx,
                bool(steps),
                bool(expected),
                bool(design_desc),
            )
            continue

        _case_id_counter += 1
        created = CaseOut(
            id=_case_id_counter,
            name=name,
            code=code,
            precondition=precondition,
            steps=steps,
            expected=expected,
        )
        _cases_store[created.id] = created
        imported.append(created)

    workbook.close()
    logger.info(
        "Excel import completed, filename=%s, scanned_rows=%d, imported=%d, skipped_missing_name=%d, skipped_missing_steps_or_expected=%d",
        file.filename,
        scanned_rows,
        len(imported),
        skipped_rows_missing_name,
        skipped_rows_missing_required_fields,
    )

    if not imported:
        logger.warning(
            "Import failed: no valid rows, filename=%s, scanned_rows=%d, headers=%s, detected_header_index=%s, fallback_offset=%d",
            file.filename,
            scanned_rows,
            normalized_headers,
            header_index,
            fallback_offset,
        )
        raise HTTPException(status_code=400, detail="No valid case rows found in the Excel file")

    return imported
